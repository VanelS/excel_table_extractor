"""
app_excel_tables.py  — v6
─────────────────────────
Détection hybride de tableaux Excel avec configuration YAML.

Passes de détection :
  0. Guidée par hints YAML (ancrage par titre, headers, row indexes)
  1. Tables Excel déclarées
  2. Hybride (header score + expansion + lookahead)
  2b. Banded rows (alternance de fill)
  3. Bordures (grilles + bordures externes)
  4. Blocs contigus résiduels

Lancement :
    pip install streamlit openpyxl pandas pyyaml
    streamlit run app_excel_tables.py
"""

from __future__ import annotations
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional
from datetime import datetime, date
import re as _re

import pandas as pd
import yaml
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  CellInfo
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@dataclass(slots=True)
class CellInfo:
    value: object = None
    bold: bool = False
    fill_key: Optional[str] = None
    font_size: float = 11.0
    border_top: bool = False
    border_bottom: bool = False
    border_left: bool = False
    border_right: bool = False
    has_any_border: bool = False
    data_type: str = "empty"

    @property
    def has_fill(self): return self.fill_key is not None
    @property
    def all_borders(self):
        return self.border_top and self.border_bottom and self.border_left and self.border_right


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  TableHint — un indice de recherche pour un tableau
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@dataclass
class TableHint:
    id: Optional[str] = None
    table_title: Optional[str] = None
    expected_headers: Optional[list[str]] = None
    expected_row_indexes: Optional[list[str]] = None
    has_total_row: Optional[bool] = None
    expected_columns: Optional[int] = None
    row_range: Optional[tuple[int, int]] = None

    @classmethod
    def from_dict(cls, d: dict):
        rr = d.get("row_range")
        return cls(
            id=d.get("id"),
            table_title=d.get("table_title"),
            expected_headers=d.get("expected_headers"),
            expected_row_indexes=d.get("expected_row_indexes"),
            has_total_row=d.get("has_total_row"),
            expected_columns=d.get("expected_columns"),
            row_range=tuple(rr) if rr else None,
        )

    @property
    def display_id(self) -> str:
        return self.id or self.table_title or "unnamed"


@dataclass
class SheetConfig:
    hints: list[TableHint] = field(default_factory=list)
    strict: bool = False


def load_yaml_config(content: str) -> dict[str, SheetConfig]:
    raw = yaml.safe_load(content) or {}
    strict = raw.get("strict", False)
    configs: dict[str, SheetConfig] = {}
    for sheet_name, sheet_data in raw.get("sheets", {}).items():
        hints = [TableHint.from_dict(t) for t in sheet_data.get("tables", [])]
        configs[sheet_name] = SheetConfig(hints=hints, strict=strict)
    return configs


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  DetectedTable
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@dataclass
class DetectedTable:
    sheet: str
    title: str
    top_left: str
    bottom_right: str
    num_rows: int
    num_cols: int
    headers: list[str] = field(default_factory=list)
    source: str = ""
    score: float = 0.0
    has_header_fill: bool = False
    has_total_row: bool = False
    has_grid_borders: bool = False
    expanded_left: bool = False
    expanded_right: bool = False
    has_empty_rows: bool = False
    matched_rules: list[str] = field(default_factory=list)

    @property
    def range_str(self): return f"{self.top_left}:{self.bottom_right}"
    @property
    def badge(self):
        return {
            "hint_guided":      "🎯 Guidé YAML",
            "excel_table":      "📊 Table Excel",
            "hybrid_detected":  "🔍 Hybride",
            "grid_detected":    "🔲 Grille",
            "contiguous_block": "📦 Bloc",
        }.get(self.source, self.source)
    @property
    def confidence(self):
        if self.score >= 70: return "haute"
        if self.score >= 40: return "moyenne"
        return "basse"


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  Fuzzy text matching
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _normalize(s: str) -> str:
    if not s: return ""
    return _re.sub(r'\s+', ' ', str(s).strip().lower())

def _fuzzy_match(needle: str, haystack: str) -> bool:
    return _normalize(needle) in _normalize(haystack) or _normalize(haystack) in _normalize(needle)

def _match_ratio(expected: list[str], actual: list) -> float:
    if not expected: return 0
    actual_norm = [_normalize(str(v)) for v in actual if v is not None]
    matched = sum(1 for e in expected
                  if any(_fuzzy_match(e, a) for a in actual_norm))
    return matched / len(expected)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  ExcelTableDetector v6
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

LOOKAHEAD = 2

class ExcelTableDetector:

    def __init__(self, filepath: str):
        self.filepath = Path(filepath)
        wb_light = load_workbook(filepath, read_only=True, data_only=True)
        self.sheetnames = wb_light.sheetnames
        wb_light.close()
        self._wb = None
        self._wb_values = None

    def _ensure_wb(self):
        if self._wb is None:
            self._wb = load_workbook(str(self.filepath), data_only=False)
        if self._wb_values is None:
            self._wb_values = load_workbook(str(self.filepath), data_only=True)

    def _build_matrix(self, ws) -> dict[tuple[int,int], CellInfo]:
        matrix = {}
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 1,
                                min_col=1, max_col=ws.max_column or 1):
            for cell in row:
                r, c = cell.row, cell.column
                fk = None
                f = cell.fill
                if f and f.patternType and f.patternType != "none" and f.start_color:
                    sc = f.start_color
                    if sc.type == "rgb" and sc.rgb: fk = f"rgb:{sc.rgb}"
                    elif sc.type == "theme": fk = f"theme:{sc.theme}:{round(sc.tint or 0,4)}"
                    elif sc.type == "indexed": fk = f"idx:{sc.value}"
                b = cell.border
                bt = b.top.style is not None if b and b.top else False
                bb = b.bottom.style is not None if b and b.bottom else False
                bl = b.left.style is not None if b and b.left else False
                br_ = b.right.style is not None if b and b.right else False
                v = cell.value
                if v is None: dt = "empty"
                elif isinstance(v, str) and v.startswith("="): dt = "formula"
                elif isinstance(v, (int, float)): dt = "number"
                elif isinstance(v, (datetime, date)): dt = "date"
                else: dt = "text"
                info = CellInfo(value=v, bold=bool(cell.font and cell.font.bold),
                    fill_key=fk, font_size=float(cell.font.size) if cell.font and cell.font.size else 11.0,
                    border_top=bt, border_bottom=bb, border_left=bl, border_right=br_,
                    has_any_border=(bt or bb or bl or br_), data_type=dt)
                if v is not None or fk is not None or info.has_any_border:
                    matrix[(r, c)] = info
        for m in ws.merged_cells.ranges:
            top = matrix.get((m.min_row, m.min_col))
            for mr in range(m.min_row, m.max_row + 1):
                for mc in range(m.min_col, m.max_col + 1):
                    if (mr, mc) not in matrix and top:
                        matrix[(mr, mc)] = CellInfo(value=top.value, bold=top.bold,
                            fill_key=top.fill_key, font_size=top.font_size, data_type=top.data_type)
        return matrix

    def _get(self, matrix, r, c) -> CellInfo:
        return matrix.get((r, c), CellInfo())

    # ────────────────────────────────────────────
    #  API publique
    # ────────────────────────────────────────────
    def detect_sheet(self, sheet_name: str,
                     config: Optional[SheetConfig] = None) -> list[DetectedTable]:
        self._ensure_wb()
        ws = self._wb[sheet_name]
        matrix = self._build_matrix(ws)
        results: list[DetectedTable] = []
        max_r = ws.max_row or 1
        max_c = ws.max_column or 1
        content_cells = {pos for pos, ci in matrix.items() if ci.value is not None}
        hints = config.hints if config else []

        # Passe 0 : Détection guidée par hints YAML
        covered = set()
        if hints:
            covered = self._hint_guided(matrix, sheet_name, results,
                                        hints, max_r, max_c, ws)

        # Passe 1 : Tables Excel déclarées
        covered |= self._declared_tables(ws, sheet_name, results, matrix, covered)

        # Passe 2 : Hybride
        covered |= self._hybrid_tables(matrix, sheet_name, results, covered,
                                        max_r, max_c, ws)
        # Passe 2b : Banded rows
        covered |= self._banded_tables(matrix, sheet_name, results, covered,
                                        max_r, max_c, ws)
        # Passe 3 : Bordures
        covered |= self._grid_tables(matrix, sheet_name, results, covered,
                                      max_r, max_c, ws)
        # Passe 4 : Blocs résiduels
        if not (config and config.strict):
            self._residual_blocks(matrix, sheet_name, results, covered,
                                  content_cells, max_r, max_c, ws)

        # Post-traitement : matcher les hints restants aux résultats
        if hints:
            self._post_match_hints(results, hints, matrix)

        # Mode strict : ne garder que les tables matchées
        if config and config.strict:
            results = [t for t in results if t.matched_rules]

        return results

    def load_table(self, table: DetectedTable) -> pd.DataFrame:
        self._ensure_wb()
        ws_val = self._wb_values[table.sheet]
        mc, mr, xc, xr = range_boundaries(f"{table.top_left}:{table.bottom_right}")
        data = []
        for row in ws_val.iter_rows(min_row=mr, max_row=xr, min_col=mc, max_col=xc):
            data.append([cell.value for cell in row])
        if not data: return pd.DataFrame()
        raw = data[0]
        headers, seen = [], {}
        for h in raw:
            name = str(h).strip() if h is not None else "Sans titre"
            name = name.replace("\n", " ")
            if name in seen: seen[name] += 1; name = f"{name}_{seen[name]}"
            else: seen[name] = 0
            headers.append(name)
        return pd.DataFrame(data[1:], columns=headers)

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    #  Passe 0 : Détection guidée par hints
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

    def _hint_guided(self, matrix, sheet, results,
                     hints: list[TableHint], max_r, max_c, ws) -> set:
        covered = set()

        for hint in hints:
            # Chercher le point d'ancrage
            anchor_row = None
            anchor_c1 = None
            anchor_c2 = None
            index_col = None

            # Stratégie A : Trouver par titre
            title_row = None
            if hint.table_title:
                title_row = self._find_title_row(matrix, hint.table_title,
                                                  max_r, max_c,
                                                  exclude_covered=covered)

            # Stratégie B : Trouver par expected_headers
            if hint.expected_headers:
                header_match = self._find_header_row(
                    matrix, hint.expected_headers, max_r, max_c,
                    search_after=title_row, exclude_covered=covered)
                if header_match:
                    anchor_row, anchor_c1, anchor_c2 = header_match

            # Stratégie C : Trouver par expected_row_indexes
            if hint.expected_row_indexes:
                idx_match = self._find_index_column(
                    matrix, hint.expected_row_indexes, max_r, max_c,
                    near_row=anchor_row or (title_row + 1 if title_row else None))
                if idx_match:
                    index_col, idx_r1, idx_r2 = idx_match
                    if anchor_row is None:
                        anchor_row = idx_r1 - 1 if idx_r1 > 1 else idx_r1

            # Si pas de header trouvé mais titre trouvé, chercher juste en dessous
            if anchor_row is None and title_row:
                for r in range(title_row + 1, min(title_row + 4, max_r + 1)):
                    row_vals = [(c, self._get(matrix, r, c))
                                for c in range(1, max_c + 1)
                                if self._get(matrix, r, c).value is not None]
                    if len(row_vals) >= 2:
                        anchor_row = r
                        anchor_c1 = row_vals[0][0]
                        anchor_c2 = row_vals[-1][0]
                        break

            if anchor_row is None:
                continue

            # Déterminer les bornes du tableau
            if anchor_c1 is None:
                row_vals = [(c, self._get(matrix, anchor_row, c))
                            for c in range(1, max_c + 1)
                            if self._get(matrix, anchor_row, c).value is not None]
                if not row_vals: continue
                anchor_c1 = row_vals[0][0]
                anchor_c2 = row_vals[-1][0]

            # Inclure la colonne d'index si trouvée et hors plage
            if index_col and index_col < anchor_c1:
                anchor_c1 = index_col
            if index_col and index_col > anchor_c2:
                anchor_c2 = index_col

            # Expansion vers le bas
            end_row = anchor_row
            total_row = None
            empty_streak = 0
            typical_fill = None

            for r in range(anchor_row + 1, max_r + 1):
                filled = sum(1 for c in range(anchor_c1, anchor_c2 + 1)
                            if self._get(matrix, r, c).value is not None)

                if filled == 0:
                    empty_streak += 1
                    if empty_streak > LOOKAHEAD:
                        break
                    continue

                if empty_streak > 0:
                    if typical_fill and abs(filled - typical_fill) > 1:
                        break
                    empty_streak = 0

                if typical_fill is None: typical_fill = filled
                else: typical_fill = round((typical_fill + filled) / 2)

                # Total row ?
                row_cells = [self._get(matrix, r, c)
                             for c in range(anchor_c1, anchor_c2 + 1)]
                non_empty = [ci for ci in row_cells if ci.value is not None]
                all_bold = non_empty and all(ci.bold for ci in non_empty)
                any_top = any(ci.border_top for ci in row_cells)

                if all_bold and any_top:
                    total_row = r
                    if hint.has_total_row:
                        end_row = r
                    break

                end_row = r

            if hint.has_total_row and total_row:
                end_row = total_row

            # Si has_total_row mais pas encore trouvé → chercher au-delà
            if hint.has_total_row and not total_row:
                for r in range(end_row + 1, min(end_row + 5, max_r + 1)):
                    row_cells = [self._get(matrix, r, c)
                                 for c in range(anchor_c1, anchor_c2 + 1)]
                    non_empty = [ci for ci in row_cells if ci.value is not None]
                    if not non_empty: continue
                    all_bold = all(ci.bold for ci in non_empty)
                    any_top = any(ci.border_top for ci in row_cells)
                    # Total = bold + border_top OU première cellule = "Total"
                    first_val = non_empty[0].value if non_empty else None
                    is_total_kw = (first_val and isinstance(first_val, str)
                                   and _normalize(first_val) in
                                   {"total","totaux","sous-total","subtotal","sum","somme"})
                    if (all_bold and any_top) or is_total_kw:
                        total_row = r
                        end_row = r
                        break

            # Contraintes row_range
            num_rows = end_row - anchor_row
            if hint.row_range:
                if num_rows < hint.row_range[0] or num_rows > hint.row_range[1]:
                    continue

            # Contrainte expected_columns
            num_cols = anchor_c2 - anchor_c1 + 1
            if hint.expected_columns and num_cols != hint.expected_columns:
                continue

            # Vérifier pas de chevauchement massif
            cells_set = set()
            for r in range(anchor_row, end_row + 1):
                for c in range(anchor_c1, anchor_c2 + 1):
                    if (r, c) not in covered:
                        cells_set.add((r, c))

            # Score
            score = 50  # base hint
            if hint.expected_headers:
                hdr_vals = [self._get(matrix, anchor_row, c).value
                            for c in range(anchor_c1, anchor_c2 + 1)]
                ratio = _match_ratio(hint.expected_headers, hdr_vals)
                score += 25 * ratio
            if hint.expected_row_indexes and index_col:
                idx_vals = [self._get(matrix, r, index_col).value
                            for r in range(anchor_row + 1, end_row + 1)]
                ratio = _match_ratio(hint.expected_row_indexes, idx_vals)
                score += 15 * ratio
            if hint.table_title and title_row:
                score += 10
            score = min(score, 100)

            # Headers
            headers = [str(self._get(matrix, anchor_row, c).value or "")
                       for c in range(anchor_c1, anchor_c2 + 1)]

            title = hint.table_title or f"Hint_{get_column_letter(anchor_c1)}{anchor_row}"

            results.append(DetectedTable(
                sheet=sheet, title=title,
                top_left=f"{get_column_letter(anchor_c1)}{anchor_row}",
                bottom_right=f"{get_column_letter(anchor_c2)}{end_row}",
                num_rows=num_rows, num_cols=num_cols,
                headers=headers, source="hint_guided", score=score,
                has_total_row=total_row is not None,
                matched_rules=[hint.display_id],
            ))
            covered |= cells_set

        return covered

    def _find_title_row(self, matrix, title: str, max_r, max_c,
                        exclude_covered: Optional[set] = None) -> Optional[int]:
        needle = _normalize(title)
        excl = exclude_covered or set()
        # Passe 1 : match exact (après normalisation)
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                if (r, c) in excl: continue
                ci = self._get(matrix, r, c)
                if ci.data_type == "formula": continue
                if ci.value and _normalize(str(ci.value)) == needle:
                    return r
        # Passe 2 : match partiel mais seulement sur des cellules texte courtes
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                if (r, c) in excl: continue
                ci = self._get(matrix, r, c)
                if ci.data_type != "text": continue
                v = str(ci.value).strip()
                if len(v) > 50: continue
                if needle in _normalize(v) and len(v) < len(needle) * 3:
                    return r
        return None

    def _find_header_row(self, matrix, expected: list[str], max_r, max_c,
                         search_after: Optional[int] = None,
                         exclude_covered: Optional[set] = None) -> Optional[tuple]:
        start = (search_after + 1) if search_after else 1
        end = min(start + 15, max_r + 1) if search_after else max_r + 1
        best = None
        best_ratio = 0
        excl = exclude_covered or set()

        for r in range(start, end):
            # Découper la ligne en segments contigus de cellules non-vides
            segments = []
            seg_start = None
            seg_vals = []
            for c in range(1, max_c + 2):
                ci = self._get(matrix, r, c) if c <= max_c else CellInfo()
                has_val = ci.value is not None and (r, c) not in excl and c <= max_c
                if has_val:
                    if seg_start is None: seg_start = c
                    seg_vals.append(str(ci.value))
                else:
                    if seg_start is not None and len(seg_vals) >= 2:
                        segments.append((seg_start, c - 1, seg_vals[:]))
                    seg_start = None
                    seg_vals = []

            # Évaluer chaque segment
            for c1, c2, vals in segments:
                ratio = _match_ratio(expected, vals)
                if ratio > best_ratio and ratio >= 0.5:
                    best_ratio = ratio
                    best = (r, c1, c2)

        return best

    def _find_index_column(self, matrix, expected: list[str], max_r, max_c,
                           near_row: Optional[int] = None) -> Optional[tuple]:
        best = None
        best_ratio = 0

        for c in range(1, max_c + 1):
            col_vals = []
            first_r = None
            last_r = None
            search_start = max(1, (near_row - 2) if near_row else 1)
            search_end = min(max_r, (near_row + 50) if near_row else max_r)

            for r in range(search_start, search_end + 1):
                ci = self._get(matrix, r, c)
                if ci.value is not None and ci.data_type == "text":
                    col_vals.append(str(ci.value))
                    if first_r is None: first_r = r
                    last_r = r

            if not col_vals or first_r is None:
                continue

            ratio = _match_ratio(expected, col_vals)
            if ratio > best_ratio and ratio >= 0.4:
                best_ratio = ratio
                best = (c, first_r, last_r)

        return best

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    #  Post-match : associer hints aux résultats existants
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

    def _post_match_hints(self, results: list[DetectedTable],
                          hints: list[TableHint], matrix):
        # Collecter les IDs déjà matchés en passe 0
        already_matched = set()
        for t in results:
            already_matched.update(t.matched_rules)

        for hint in hints:
            hint_id = hint.display_id
            if hint_id in already_matched:
                continue

            best_table = None
            best_score = 0

            for t in results:
                match_score = 0

                if hint.table_title and _fuzzy_match(hint.table_title, t.title):
                    match_score += 30

                if hint.expected_headers and t.headers:
                    ratio = _match_ratio(hint.expected_headers, t.headers)
                    match_score += 40 * ratio

                if hint.expected_columns and t.num_cols == hint.expected_columns:
                    match_score += 10

                if match_score > best_score:
                    best_score = match_score
                    best_table = t

            if best_table and best_score >= 20:
                if hint_id not in best_table.matched_rules:
                    best_table.matched_rules.append(hint_id)
                best_table.score = min(best_table.score + 15, 100)

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    #  Passe 1 : Tables Excel déclarées
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

    def _declared_tables(self, ws, sheet, results, matrix, already_covered) -> set:
        covered = set()
        for tbl in ws.tables.values():
            min_c, min_r, max_c, max_r = range_boundaries(tbl.ref)
            cells = set()
            for r in range(min_r, max_r + 1):
                for c in range(min_c, max_c + 1):
                    cells.add((r, c))

            if cells & already_covered:
                continue

            has_total = False
            tr = max_r + 1
            ci_t = self._get(matrix, tr, min_c)
            if ci_t.value is not None and ci_t.bold and ci_t.border_top:
                has_total = True
                max_r = tr
                for c in range(min_c, max_c + 1): cells.add((tr, c))

            title = self._find_section_title(matrix, min_r, min_c, max_c, ws)
            if not title: title = tbl.displayName
            headers = [str(self._get(matrix, min_r, c).value or "")
                       for c in range(min_c, max_c + 1)]
            hdr_fill = self._get(matrix, min_r, min_c).fill_key

            results.append(DetectedTable(
                sheet=sheet, title=title,
                top_left=f"{get_column_letter(min_c)}{min_r}",
                bottom_right=f"{get_column_letter(max_c)}{max_r}",
                num_rows=max_r - min_r, num_cols=max_c - min_c + 1,
                headers=headers, source="excel_table", score=100,
                has_header_fill=hdr_fill is not None, has_total_row=has_total))
            covered |= cells
        return covered

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    #  Header score (avec gardes anti-total/titre)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

    def _header_score(self, matrix, row, c1, c2, max_r) -> float:
        score = 0.0
        cells = [self._get(matrix, row, c) for c in range(c1, c2 + 1)]
        non_empty = [ci for ci in cells if ci.value is not None]
        if not non_empty: return 0

        all_bold = all(ci.bold for ci in non_empty)
        any_top = any(ci.border_top for ci in cells)
        if all_bold and any_top and row > 1:
            prev = [self._get(matrix, row-1, c) for c in range(c1, c2+1)]
            prev_ne = [ci for ci in prev if ci.value is not None]
            if prev_ne and not all(ci.bold for ci in prev_ne): return -5

        total_kw = {"total","totaux","sous-total","subtotal","sum","somme","grand total","net","solde"}
        fv = non_empty[0].value
        if fv and isinstance(fv, str) and _normalize(fv) in total_kw and any_top: return -5
        if non_empty and all(ci.font_size >= 16 for ci in non_empty):
            if all(ci.data_type == "text" for ci in non_empty): return -3

        types_in = set(ci.data_type for ci in non_empty)
        if "number" in types_in and "text" not in types_in: return -5

        bold_r = sum(1 for ci in non_empty if ci.bold) / len(non_empty)
        if bold_r >= 0.8: score += 3
        elif bold_r >= 0.5: score += 1.5

        fills = set(ci.fill_key for ci in non_empty if ci.fill_key)
        if fills:
            nf = set(self._get(matrix, row+1, c).fill_key for c in range(c1, c2+1)
                     if self._get(matrix, row+1, c).fill_key)
            if fills != nf: score += 3

        bot_r = sum(1 for ci in non_empty if ci.border_bottom) / len(non_empty)
        if bot_r >= 0.5: score += 2

        if types_in <= {"text"} and row + 1 <= max_r:
            nc = [self._get(matrix, row+1, c) for c in range(c1, c2+1)]
            nn = [ci for ci in nc if ci.value is not None]
            if nn and set(ci.data_type for ci in nn) & {"number","date","formula"}:
                score += 5

        if row + 2 <= max_r:
            bs = self._banded_rows_score(matrix, row+1, c1, c2, min(row+6, max_r))
            if bs >= 2: score += 3

        return score

    def _banded_rows_score(self, matrix, start_row, c1, c2, end_row) -> float:
        row_fills = []
        for r in range(start_row, end_row + 1):
            fills = set()
            for c in range(c1, c2+1):
                fk = self._get(matrix, r, c).fill_key
                if fk: fills.add(fk)
            if len(fills) == 1: row_fills.append(fills.pop())
            elif len(fills) == 0: row_fills.append(None)
            else: return 0
        if len(row_fills) < 3: return 0
        distinct = set(f for f in row_fills if f is not None)
        if len(distinct) != 2: return 0
        return sum(1 for i in range(1, len(row_fills))
                   if row_fills[i] != row_fills[i-1] and row_fills[i] is not None)

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    #  Passe 2 : Hybride
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

    def _hybrid_tables(self, matrix, sheet, results, covered, max_r, max_c, ws) -> set:
        new_covered = set()
        THRESHOLD = 3.0
        header_runs = []

        for r in range(1, max_r + 1):
            segments = []
            seg_start = None
            for c in range(1, max_c + 2):
                ci = self._get(matrix, r, c) if c <= max_c else CellInfo()
                in_cov = (r,c) in covered or (r,c) in new_covered
                has_val = ci.value is not None and not in_cov and c <= max_c
                if has_val:
                    if seg_start is None: seg_start = c
                else:
                    if seg_start is not None and c - seg_start >= 2:
                        segments.append((seg_start, c-1))
                    seg_start = None
            for c1, c2 in segments:
                hs = self._header_score(matrix, r, c1, c2, max_r)
                if hs >= THRESHOLD:
                    fk = self._get(matrix, r, c1).fill_key
                    header_runs.append((r, c1, c2, fk, hs))

        header_runs.sort(key=lambda x: -x[4])
        used_cells = set()

        for hdr_row, style_c1, style_c2, fk, hs in header_runs:
            if any((hdr_row,c) in used_cells or (hdr_row,c) in covered
                   for c in range(style_c1, style_c2+1)):
                continue

            last_data_row = hdr_row
            total_row = None
            empty_streak = 0
            has_empty_rows = False
            typical_fill = None
            r = hdr_row + 1
            while r <= max_r:
                filled = sum(1 for c in range(style_c1, style_c2+1)
                            if self._get(matrix, r, c).value is not None)
                if filled == 0:
                    empty_streak += 1
                    if empty_streak > LOOKAHEAD: break
                    r += 1; continue
                if empty_streak > 0:
                    if typical_fill and abs(filled - typical_fill) > 1: break
                    has_empty_rows = True; empty_streak = 0
                if typical_fill is None: typical_fill = filled
                else: typical_fill = round((typical_fill + filled) / 2)
                rc = [self._get(matrix, r, c) for c in range(style_c1, style_c2+1)]
                ne = [ci for ci in rc if ci.value is not None]
                if ne and all(ci.bold for ci in ne) and any(ci.border_top for ci in rc):
                    total_row = r; break
                last_data_row = r; r += 1

            end_row = total_row or last_data_row
            if end_row - hdr_row < 1: continue

            final_c1, final_c2 = style_c1, style_c2
            exp_l = exp_r = False
            c = style_c1 - 1
            while c >= 1:
                cf = sum(1 for r2 in range(hdr_row, end_row+1)
                         if self._get(matrix, r2, c).value is not None and (r2,c) not in covered)
                if cf >= max((end_row-hdr_row+1)*0.3, 1): final_c1 = c; exp_l = True; c -= 1
                else: break
            c = style_c2 + 1
            while c <= max_c:
                cf = sum(1 for r2 in range(hdr_row, end_row+1)
                         if self._get(matrix, r2, c).value is not None and (r2,c) not in covered)
                if cf >= max((end_row-hdr_row+1)*0.3, 1): final_c2 = c; exp_r = True; c += 1
                else: break

            num_cols = final_c2 - final_c1 + 1
            num_rows = end_row - hdr_row
            cells_set = set()
            ov = 0; tot = 0
            for r2 in range(hdr_row, end_row+1):
                for c2 in range(final_c1, final_c2+1):
                    tot += 1
                    if (r2,c2) in covered or (r2,c2) in new_covered: ov += 1
                    else: cells_set.add((r2,c2))
            if tot > 0 and ov / tot > 0.3: continue

            score = min(hs*5, 30)
            score += 15 if num_cols >= 3 else 8 if num_cols >= 2 else 0
            score += 15 if num_rows >= 3 else 8 if num_rows >= 1 else 0
            score += 10 if total_row else 0
            score += self._type_consistency_score(matrix, hdr_row, end_row, final_c1, final_c2)
            be = (total_row-1) if total_row else end_row
            bc = max((be-hdr_row)*num_cols, 1)
            fc = sum(1 for r2 in range(hdr_row+1, be+1) for c2 in range(final_c1, final_c2+1)
                     if self._get(matrix, r2, c2).value is not None)
            d = fc / bc
            score += 10 if d >= 0.4 else 5 if d >= 0.2 else 0
            if exp_l or exp_r: score += 5
            score = min(score, 95)

            title = self._find_section_title(matrix, hdr_row, final_c1, final_c2, ws)
            if not title: title = f"Tableau_{get_column_letter(final_c1)}{hdr_row}"
            headers = [str(self._get(matrix, hdr_row, c).value or "") for c in range(final_c1, final_c2+1)]

            results.append(DetectedTable(
                sheet=sheet, title=title,
                top_left=f"{get_column_letter(final_c1)}{hdr_row}",
                bottom_right=f"{get_column_letter(final_c2)}{end_row}",
                num_rows=num_rows, num_cols=num_cols, headers=headers,
                source="hybrid_detected", score=score,
                has_header_fill=fk is not None, has_total_row=total_row is not None,
                expanded_left=exp_l, expanded_right=exp_r, has_empty_rows=has_empty_rows))
            new_covered |= cells_set
            for r2 in range(hdr_row, end_row+1):
                for c2 in range(final_c1, final_c2+1): used_cells.add((r2, c2))

        return new_covered

    def _type_consistency_score(self, matrix, hdr_row, end_row, c1, c2) -> float:
        if end_row - hdr_row < 2: return 0
        cons = 0; total = 0
        for c in range(c1, c2+1):
            types = [self._get(matrix, r, c).data_type
                     for r in range(hdr_row+1, end_row+1)
                     if self._get(matrix, r, c).value is not None]
            if len(types) >= 2:
                total += 1
                dom = max(set(types), key=types.count)
                if types.count(dom)/len(types) >= 0.7: cons += 1
        return 10 * (cons / total) if total else 0

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    #  Passe 2b : Banded rows
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

    def _banded_tables(self, matrix, sheet, results, covered, max_r, max_c, ws) -> set:
        new_covered = set()
        for c_start in range(1, max_c+1):
            for r_start in range(1, max_r-2):
                if (r_start, c_start) in covered or (r_start, c_start) in new_covered: continue
                row_fills = []; r_end = r_start
                for r in range(r_start, max_r+1):
                    ci = self._get(matrix, r, c_start)
                    if ci.value is None and ci.fill_key is None: break
                    row_fills.append(ci.fill_key); r_end = r
                if len(row_fills) < 4: continue
                distinct = set(f for f in row_fills if f is not None)
                if len(distinct) != 2: continue
                alt = sum(1 for i in range(1,len(row_fills))
                          if row_fills[i] != row_fills[i-1] and row_fills[i] is not None and row_fills[i-1] is not None)
                if alt < 2: continue
                c_end = c_start
                for c in range(c_start+1, max_c+1):
                    cf = [self._get(matrix,r,c).fill_key for r in range(r_start, r_end+1)]
                    if cf == row_fills: c_end = c
                    else: break
                nc = c_end - c_start + 1; nr = r_end - r_start
                if nc < 2 or nr < 3: continue
                cs = set(); ov = 0
                for r in range(r_start, r_end+1):
                    for c in range(c_start, c_end+1):
                        if (r,c) in covered or (r,c) in new_covered: ov += 1
                        else: cs.add((r,c))
                tc = (r_end-r_start+1)*nc
                if tc > 0 and ov/tc > 0.3: continue
                hdr = None
                if r_start > 1:
                    pc = self._get(matrix, r_start-1, c_start)
                    if pc.value is not None and pc.fill_key not in distinct:
                        hdr = r_start - 1
                        for c in range(c_start, c_end+1): cs.add((hdr, c))
                actual = hdr or r_start
                score = min(35 + (15 if nc >= 3 else 8) + (15 if nr >= 3 else 8) +
                            self._type_consistency_score(matrix, actual, r_end, c_start, c_end), 90)
                title = self._find_section_title(matrix, actual, c_start, c_end, ws)
                if not title: title = f"Banded_{get_column_letter(c_start)}{actual}"
                headers = [str(self._get(matrix, actual, c).value or "") for c in range(c_start, c_end+1)]
                results.append(DetectedTable(sheet=sheet, title=title,
                    top_left=f"{get_column_letter(c_start)}{actual}",
                    bottom_right=f"{get_column_letter(c_end)}{r_end}",
                    num_rows=r_end-actual, num_cols=nc, headers=headers,
                    source="hybrid_detected", score=score, has_header_fill=hdr is not None))
                new_covered |= cs; break
        return new_covered

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    #  Passe 3 : Bordures
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

    def _grid_tables(self, matrix, sheet, results, covered, max_r, max_c, ws) -> set:
        nc = set()
        bordered = {p for p, ci in matrix.items() if ci.all_borders and p not in covered}
        if bordered:
            for block in self._flood_fill(bordered, gap=0):
                self._score_grid(block, matrix, sheet, results, nc, ws)
        bcells = {p for p, ci in matrix.items() if ci.has_any_border and p not in covered and p not in nc}
        if bcells:
            for block in self._detect_ext_rects(matrix, bcells, covered | nc, max_r, max_c):
                self._score_grid(block, matrix, sheet, results, nc, ws)
        return nc

    def _detect_ext_rects(self, matrix, bcells, covered, max_r, max_c):
        rects = []
        corners = [(r,c) for (r,c), ci in matrix.items()
                    if ci.border_top and ci.border_left and (r,c) not in covered]
        for r1, c1 in corners:
            c2 = None
            for c in range(c1+1, min(c1+20, max_c+1)):
                ci = self._get(matrix, r1, c)
                if ci.border_top and ci.border_right: c2 = c; break
                if not ci.border_top and not ci.has_any_border and ci.value is None: break
            if not c2 or c2-c1 < 1: continue
            r2 = None
            for r in range(r1+1, min(r1+50, max_r+1)):
                ci = self._get(matrix, r, c1)
                if ci.border_bottom and ci.border_left: r2 = r; break
                if not ci.border_left and not ci.has_any_border and ci.value is None: break
            if not r2 or r2-r1 < 1: continue
            ci_br = self._get(matrix, r2, c2)
            if not (ci_br.border_bottom and ci_br.border_right): continue
            filled = sum(1 for r in range(r1,r2+1) for c in range(c1,c2+1)
                         if self._get(matrix,r,c).value is not None)
            total = (r2-r1+1)*(c2-c1+1)
            if filled < total*0.3: continue
            ov = sum(1 for r in range(r1,r2+1) for c in range(c1,c2+1) if (r,c) in covered)
            if ov > total*0.3: continue
            rects.append({(r,c) for r in range(r1,r2+1) for c in range(c1,c2+1)})
        return rects

    def _score_grid(self, block, matrix, sheet, results, nc, ws):
        rows_b = [r for r,c in block]; cols_b = [c for r,c in block]
        r1,r2,c1,c2 = min(rows_b),max(rows_b),min(cols_b),max(cols_b)
        nr = r2-r1; ncols = c2-c1+1
        if nr < 1 or ncols < 2: return
        score = 30 + (15 if ncols >= 3 else 8) + (15 if nr >= 3 else 8)
        score += self._type_consistency_score(matrix, r1, r2, c1, c2)
        filled = sum(1 for r,c in block if self._get(matrix,r,c).value is not None)
        score += 10 if filled/max(len(block),1) >= 0.5 else 5 if filled/max(len(block),1) >= 0.3 else 0
        fv = [self._get(matrix,r1,c) for c in range(c1,c2+1) if self._get(matrix,r1,c).value is not None]
        if fv and all(ci.bold for ci in fv): score += 5
        score = min(score, 90)
        title = self._find_section_title(matrix, r1, c1, c2, ws) or f"Grille_{get_column_letter(c1)}{r1}"
        headers = [str(self._get(matrix,r1,c).value or "") for c in range(c1,c2+1)]
        results.append(DetectedTable(sheet=sheet, title=title,
            top_left=f"{get_column_letter(c1)}{r1}", bottom_right=f"{get_column_letter(c2)}{r2}",
            num_rows=nr, num_cols=ncols, headers=headers,
            source="grid_detected", score=score, has_grid_borders=True))
        nc.update((r,c) for r in range(r1,r2+1) for c in range(c1,c2+1))

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    #  Passe 4 : Blocs résiduels
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

    def _residual_blocks(self, matrix, sheet, results, covered, content, max_r, max_c, ws):
        remaining = {p for p in content if p not in covered}
        for block in self._flood_fill(remaining, gap=0):
            rows_b = [r for r,c in block]; cols_b = [c for r,c in block]
            r1,r2,c1,c2 = min(rows_b),max(rows_b),min(cols_b),max(cols_b)
            nr = r2-r1; nc = c2-c1+1
            if nr < 1 and nc < 2: continue
            score = 15 if (nr >= 3 and nc >= 2) else 8 if nr >= 1 else 0
            fv = [self._get(matrix,r1,c) for c in range(c1,c2+1) if self._get(matrix,r1,c).value is not None]
            if fv and all(ci.bold for ci in fv): score += 10
            fills = set(self._get(matrix,r1,c).fill_key for c in range(c1,c2+1) if self._get(matrix,r1,c).fill_key)
            if len(fills) == 1: score += 8
            tc = max((r2-r1+1)*nc,1); dens = len(block)/tc
            score += 8 if dens >= 0.4 else 4 if dens >= 0.2 else 0
            score += self._type_consistency_score(matrix, r1, r2, c1, c2)
            has_num = any(self._get(matrix,r,c).data_type in ("number","date") for r in range(r1,r2+1) for c in range(c1,c2+1))
            fs = [self._get(matrix,r,c).font_size for r in range(r1,r2+1) for c in range(c1,c2+1) if self._get(matrix,r,c).value is not None]
            if fs and min(fs) >= 16: score = max(score-20, 3)
            if nc == 1:
                av = [self._get(matrix,r,c1) for r in range(r1,r2+1) if self._get(matrix,r,c1).value is not None]
                if av and all(ci.bold for ci in av) and all(ci.data_type == "text" for ci in av):
                    score = max(score-25, 3)
            av2 = [self._get(matrix,r,c) for r in range(r1,r2+1) for c in range(c1,c2+1) if self._get(matrix,r,c).value is not None]
            if av2 and all(ci.bold for ci in av2) and all(ci.data_type == "text" for ci in av2) and not has_num:
                score = max(score-15, 3)
            score = min(score, 55)
            title = self._find_section_title(matrix, r1, c1, c2, ws) or f"Bloc_{get_column_letter(c1)}{r1}"
            headers = [str(self._get(matrix,r1,c).value or "") for c in range(c1,c2+1) if self._get(matrix,r1,c).value is not None]
            results.append(DetectedTable(sheet=sheet, title=title,
                top_left=f"{get_column_letter(c1)}{r1}", bottom_right=f"{get_column_letter(c2)}{r2}",
                num_rows=nr, num_cols=nc, headers=headers, source="contiguous_block", score=score))

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    #  Utilitaires
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

    def _find_section_title(self, matrix, top_row, c1, c2, ws) -> Optional[str]:
        for m in ws.merged_cells.ranges:
            if m.min_row >= top_row-2 and m.max_row < top_row:
                if m.min_col <= c2 and m.max_col >= c1:
                    v = self._get(matrix, m.min_row, m.min_col).value
                    if v and isinstance(v, str): return str(v).strip().replace("\n"," ")
        for off in [1, 2]:
            r = top_row - off
            if r < 1: break
            for c in range(c1, c2+1):
                ci = self._get(matrix, r, c)
                if ci.value and isinstance(ci.value, str) and ci.bold and ci.font_size >= 14:
                    return str(ci.value).strip().replace("\n"," ")
        if top_row > 1:
            ci = self._get(matrix, top_row-1, c1)
            if ci.value and isinstance(ci.value, str) and len(str(ci.value).strip()) > 1:
                return str(ci.value).strip()
        return None

    @staticmethod
    def _flood_fill(cells, gap=0):
        remaining, blocks = set(cells), []
        while remaining:
            seed = next(iter(remaining))
            block, queue = set(), [seed]
            while queue:
                cur = queue.pop()
                if cur in block or cur not in remaining: continue
                block.add(cur); remaining.discard(cur)
                r, c = cur
                for dr in range(-gap-1, gap+2):
                    for dc in range(-gap-1, gap+2):
                        if dr == 0 and dc == 0: continue
                        nb = (r+dr, c+dc)
                        if nb in remaining: queue.append(nb)
            if block: blocks.append(block)
        return blocks


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  Interface Streamlit
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

st.set_page_config(page_title="Excel Table Detector", page_icon="📊", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:ital,wght@0,400;0,500;0,700&family=JetBrains+Mono:wght@400;500&display=swap');
html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
h1,h2,h3 { font-family: 'DM Sans', sans-serif; font-weight: 700; }
.header-bar { background: linear-gradient(135deg, #0f172a 0%, #1e293b 50%, #334155 100%);
  padding: 2rem 2.5rem; border-radius: 16px; margin-bottom: 1.5rem; color: white; }
.header-bar h1 { color: white; margin: 0 0 .3rem 0; font-size: 1.8rem; }
.header-bar p { color: #94a3b8; margin: 0; font-size: .9rem; line-height: 1.5; }
.stat-row { display:flex; gap:.8rem; margin-bottom:1.5rem; flex-wrap:wrap; }
.stat-card { flex:1; min-width:95px; background:#f8fafc; border:1px solid #e2e8f0;
  border-radius:12px; padding:.8rem 1rem; text-align:center; }
.stat-card .num { font-size:1.5rem; font-weight:700; color:#1e293b; }
.stat-card .lbl { font-size:.7rem; color:#64748b; text-transform:uppercase; letter-spacing:.04em; margin-top:.1rem; }
.table-card { background:#fff; border:1px solid #e2e8f0; border-radius:14px;
  padding:1.2rem 1.4rem; margin-bottom:.8rem; transition:box-shadow .2s; }
.table-card:hover { box-shadow:0 4px 20px rgba(0,0,0,.06); }
.table-card .tc-header { display:flex; justify-content:space-between; align-items:center; margin-bottom:.4rem; }
.table-card .tc-title { font-weight:700; font-size:1rem; color:#1e293b; }
.badge { display:inline-block; font-size:.65rem; font-weight:600; padding:.15rem .5rem; border-radius:6px; letter-spacing:.02em; }
.badge-hint  { background:#fef3c7; color:#92400e; }
.badge-excel { background:#dcfce7; color:#166534; }
.badge-hybrid { background:#e0e7ff; color:#3730a3; }
.badge-grid  { background:#f1f5f9; color:#475569; }
.badge-block { background:#f1f5f9; color:#64748b; }
.meta { font-family:'JetBrains Mono',monospace; font-size:.74rem; color:#64748b; }
.conf-bar { display:inline-block; height:5px; border-radius:3px; vertical-align:middle; }
.conf-high { background:#22c55e; } .conf-medium { background:#f59e0b; } .conf-low { background:#ef4444; }
.sheet-chip { display:inline-block; background:#f1f5f9; border:1px solid #cbd5e1;
  border-radius:8px; padding:.22rem .65rem; font-size:.8rem; color:#475569; margin:0 .25rem .25rem 0; }
.style-pills { display:flex; gap:.35rem; margin-top:.3rem; flex-wrap:wrap; }
.pill { display:inline-block; font-size:.64rem; padding:.1rem .4rem; border-radius:4px; background:#f1f5f9; color:#475569; }
.pill-hint  { background:#fef3c7; color:#92400e; }
.pill-fill  { background:#fef3c7; color:#92400e; }
.pill-total { background:#dcfce7; color:#166534; }
.pill-expand { background:#ede9fe; color:#6d28d9; }
.pill-grid  { background:#e0e7ff; color:#3730a3; }
.pill-gap   { background:#fce7f3; color:#9d174d; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="header-bar">
    <h1>📊 Excel Table Detector v6</h1>
    <p>Détection guidée par configuration YAML — Hints par titre, headers,
       row indexes · Scoring hybride · Mode strict</p>
</div>
""", unsafe_allow_html=True)

col_up1, col_up2 = st.columns([1, 1])
with col_up1:
    uploaded = st.file_uploader("📁 Fichier Excel", type=["xlsx","xlsm"], label_visibility="collapsed")
with col_up2:
    yaml_file = st.file_uploader("📝 Config YAML (optionnel)", type=["yml","yaml"], label_visibility="collapsed")

if not uploaded:
    st.info("⬆️  Chargez un fichier **.xlsx** et optionnellement un fichier **.yml** de configuration.")
    st.stop()

tmp_path = Path("/tmp") / uploaded.name
tmp_path.write_bytes(uploaded.getvalue())
detector = ExcelTableDetector(str(tmp_path))

# Charger la config YAML
yaml_configs: dict[str, SheetConfig] = {}
if yaml_file:
    try:
        yaml_configs = load_yaml_config(yaml_file.getvalue().decode("utf-8"))
        hint_count = sum(len(sc.hints) for sc in yaml_configs.values())
        st.success(f"✅ Config chargée — {len(yaml_configs)} feuille(s), {hint_count} hint(s)")
    except Exception as e:
        st.error(f"Erreur YAML : {e}")

st.markdown(f"### 📑 {len(detector.sheetnames)} onglet(s)")
chips = " ".join(f'<span class="sheet-chip">{s}</span>' for s in detector.sheetnames)
st.markdown(chips, unsafe_allow_html=True)
st.markdown("")

selected = st.multiselect("Feuilles à analyser", detector.sheetnames, default=[],
                           placeholder="Choisissez…")
if not selected:
    st.warning("👈  Sélectionnez au moins une feuille.")
    st.stop()

tables: list[DetectedTable] = []
for s in selected:
    with st.spinner(f"Analyse de « {s} »…"):
        cfg = yaml_configs.get(s)
        tables += detector.detect_sheet(s, config=cfg)
tables.sort(key=lambda t: t.score, reverse=True)

ne = sum(1 for t in tables if t.source == "excel_table")
nh = sum(1 for t in tables if t.source in ("hybrid_detected","hint_guided"))
ng = sum(1 for t in tables if t.source == "grid_detected")
nb = sum(1 for t in tables if t.source == "contiguous_block")
nm = sum(1 for t in tables if t.matched_rules)
avg = sum(t.score for t in tables) / max(len(tables), 1)

st.markdown(f"""
<div class="stat-row">
    <div class="stat-card"><div class="num">{nm}</div><div class="lbl">Matchés YAML</div></div>
    <div class="stat-card"><div class="num">{ne}</div><div class="lbl">Tables Excel</div></div>
    <div class="stat-card"><div class="num">{nh}</div><div class="lbl">Hybrides</div></div>
    <div class="stat-card"><div class="num">{ng}</div><div class="lbl">Grilles</div></div>
    <div class="stat-card"><div class="num">{nb}</div><div class="lbl">Blocs</div></div>
    <div class="stat-card"><div class="num">{len(tables)}</div><div class="lbl">Total</div></div>
    <div class="stat-card"><div class="num">{avg:.0f}%</div><div class="lbl">Score moy.</div></div>
</div>
""", unsafe_allow_html=True)

with st.sidebar:
    st.markdown("### Filtres")
    stypes = st.multiselect("Type", ["hint_guided","excel_table","hybrid_detected","grid_detected","contiguous_block"],
        ["hint_guided","excel_table","hybrid_detected","grid_detected","contiguous_block"],
        format_func=lambda x:{"hint_guided":"🎯 YAML","excel_table":"📊 Excel","hybrid_detected":"🔍 Hybride",
                               "grid_detected":"🔲 Grille","contiguous_block":"📦 Bloc"}[x])
    ms = st.slider("Score minimum", 0, 100, 40, step=5)
    mr = st.slider("Lignes minimum", 0, 20, 0)
    if len(selected) > 1:
        fs = st.multiselect("Feuille", selected, selected)
    else: fs = selected

filtered = [t for t in tables if t.sheet in fs and t.source in stypes and t.score >= ms and t.num_rows >= mr]

st.markdown(f"### {len(filtered)} tableau(x)")
if not filtered:
    st.warning("Aucun résultat.")
    st.stop()

for i, tbl in enumerate(filtered):
    bcls = {"hint_guided":"badge-hint","excel_table":"badge-excel","hybrid_detected":"badge-hybrid",
            "grid_detected":"badge-grid","contiguous_block":"badge-block"}.get(tbl.source,"badge-block")
    ccls = {"haute":"conf-high","moyenne":"conf-medium","basse":"conf-low"}.get(tbl.confidence,"conf-low")
    bw = max(int(tbl.score*0.8), 5)
    pills = ""
    if tbl.matched_rules:
        for rule_id in tbl.matched_rules:
            pills += f'<span class="pill pill-hint">Règle: {rule_id}</span>'
    if tbl.has_header_fill: pills += '<span class="pill pill-fill">Fill en-tête</span>'
    if tbl.has_total_row: pills += '<span class="pill pill-total">Ligne total</span>'
    if tbl.has_grid_borders: pills += '<span class="pill pill-grid">Grille</span>'
    if tbl.expanded_left: pills += '<span class="pill pill-expand">← index</span>'
    if tbl.expanded_right: pills += '<span class="pill pill-expand">→ étendu</span>'
    if tbl.has_empty_rows: pills += '<span class="pill pill-gap">Gaps tolérés</span>'

    st.markdown(f"""
    <div class="table-card">
        <div class="tc-header">
            <span class="tc-title">{tbl.title}</span>
            <span class="badge {bcls}">{tbl.badge}</span>
        </div>
        <span class="meta">
            {tbl.sheet} · <b>{tbl.range_str}</b> · {tbl.num_rows}l × {tbl.num_cols}c ·
            Score <b>{tbl.score:.0f}%</b>
            <span class="conf-bar {ccls}" style="width:{bw}px"></span> ({tbl.confidence})
        </span>
        <div class="style-pills">{pills}</div>
    </div>
    """, unsafe_allow_html=True)

    with st.expander(f"🔍  {tbl.title}", expanded=False):
        df = detector.load_table(tbl)
        if df.empty: st.caption("_(vide)_")
        else:
            st.dataframe(df, use_container_width=True, hide_index=True)
            st.download_button(f"⬇️ CSV", df.to_csv(index=False).encode("utf-8"),
                file_name=f"{tbl.title}.csv", mime="text/csv", key=f"dl_{i}")
